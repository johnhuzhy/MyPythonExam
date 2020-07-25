import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime as dt


def read_excel(filename):
    min_colm = 1
    max_colm = 3
    min_row = 2
    max_row = 7
    wb = load_workbook(filename)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        print('シート名：', sheetname)
        print('行数：', len(list(ws.rows)))
        print('列数：', len(list(ws.columns)))

        area_data = np.matrix(
            np.zeros((max_row - min_row + 1, max_colm - min_colm + 1)), dtype=str)
        for col in ws.iter_cols(min_col=min_colm, max_col=max_colm, min_row=min_row, max_row=max_row):
            for cell in col:
                col_index = cell.col_idx
                row_index = cell.row
                area_data[row_index - min_row,
                          col_index - min_colm] = cell.value

        print(area_data)


def write_excel(filename, data):
    wb = load_workbook(filename)
    df = pd.DataFrame(data)
    sheet = wb.create_sheet('writetest')
    for c in dataframe_to_rows(df, index=True, header=True):
        sheet.append(c)
    
    sheet['A1'].value = dt.now()
    sheet['A1'].number_format = 'yyyy-mm-dd hh:mm:ss'
    sheet.column_dimensions['C'].number_format = u'#,##0;[Red]-#,##0'
    wb.save(filename)


if __name__ == "__main__":
    # read_excel('./prop/20200725test.xlsx')
    array = {'Nm': ['A', 'B', 'C', 'D', 'E'], 'Length': [
        167, 178, 172, 180, 175, ], 'Weight': [63, 70, 71, 78, 70]}
    write_excel('./prop/20200725test.xlsx', array)
